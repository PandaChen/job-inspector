#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
job-inspector: 重点用户职务信息巡检脚本
列映射（实际表格）：
  A列(0): 姓名
  B列(1): 所在组织机构路径（含"暂无粤政易账号"时跳过）
  C列(2): 粤政易职务信息（"/"时表示无职务记录）
  D列(3): 网上公开职务信息（比对基准）
  E列(4): 参考网址
  F列(5): 巡检结果（写入：一致/不一致/无网址/无帐号）
  G列(6): 网页爬取职务（写入，新增列）
"""

import sys
import re
import time
import random
import argparse
from typing import Optional
from urllib.parse import urlparse

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "beautifulsoup4"])
    import requests
    from bs4 import BeautifulSoup

import warnings
warnings.filterwarnings("ignore")


# ─────────────────────────────────────────────
# 职务关键词
# ─────────────────────────────────────────────

TITLE_KEYWORDS = [
    '主任', '副主任', '委员', '委员长', '副委员长',
    '省长', '副省长', '市长', '副市长', '县长', '副县长',
    '书记', '副书记', '秘书长', '副秘书长',
    '厅长', '副厅长', '局长', '副局长', '处长', '副处长',
    '部长', '副部长',
    '院长', '副院长', '所长', '校长', '副校长',
    '检察长', '副检察长', '总队长', '副总队长',
    '理事长', '副理事长', '会长', '副会长',
    '董事长', '总经理',
    '主席', '副主席',
    '成员',   # 党组成员、常委会成员等
]


# ─────────────────────────────────────────────
# 规范化工具
# ─────────────────────────────────────────────

def normalize_title(title: str, name: str = "") -> str:
    """
    规范化职务字符串，用于比对：
    1. 去中英文括号及括号内容
    2. 去"广东"/"广东省"/"省"前缀（仅开头的地名前缀）
    3. 去尾部冒号/句号等标点
    4. 去开头的"姓名："前缀（如"陈新烈：XXX"）
    5. 顿号与逗号等价（统一为顿号）
    6. 去无关后缀（逗号/顿号后面的非职务内容，如"高级指挥长消防救援衔"、"主持全面工作…"）
    7. 去空格
    """
    if not title:
        return ""

    # 1. 去中文括号及内容
    title = re.sub(r'（[^）]*）', '', title)
    # 去英文括号及内容
    title = re.sub(r'\([^)]*\)', '', title)

    title = title.strip()

    # 4. 去"姓名："开头（如"陈新烈：XXX职务"）
    if name:
        norm_n = name.replace(' ', '').replace('\u3000', '').strip()
        title = re.sub(r'^' + re.escape(norm_n) + r'[：:]\s*', '', title)

    # 2. 去开头的省份/地名前缀
    #    规则：
    #    - 「广东省委...」/「广东省人大...」/「广东省政协...」→ 只去「广东」，保留「省委...」
    #      这样两边都归一为「省委常委」，避免「广东省委常委」vs「省委常委」不一致
    #    - 「广东省政府...」/「广东省XXX厅/局/院...」→ 去「广东省」，保留机构名
    #    - 其余「广东XXX」→ 去「广东」
    title = re.sub(r'^广东(省委|省人大|省政协)', r'\1', title.strip())  # 广东省委→省委
    title = re.sub(r'^广东省', '', title.strip())                        # 广东省政府→政府
    title = re.sub(r'^广东', '', title.strip())                          # 广东xxx→xxx

    # 7. 去空格前，先处理"职务  机构名"尾部后缀（如"副厅长  广东省科学技术厅"）
    #    职务关键词后跟2个以上空格+机构名，截断
    title = re.sub(r'((?:' + '|'.join(re.escape(k) for k in TITLE_KEYWORDS) + r'))\s{2,}[^\s,，。；;：:]+$', r'\1', title)

    # 7. 去空格
    title = title.replace(' ', '').replace('\u3000', '').strip()

    # 3. 去尾部冒号/句号/分号等标点
    title = title.rstrip('：:。；;，、 \t')
    # 5. 顿号与逗号等价（统一替换为逗号便于比对）
    title = title.replace('、', ',').replace('，', ',')

    # 6. 去无关后缀：若最后一个逗号后内容不含职务关键词，则截断
    #    例："党委副书记、总队长,高级指挥长消防救援衔" → "党委副书记,总队长"
    #    例："党组书记、局长,主持全面工作,负责内部审计工作" → "党组书记,局长"
    parts = title.split(',')
    trimmed = []
    for part in parts:
        p = part.strip()
        if not p:
            continue
        # 判断是否含职务关键词
        if any(kw in p for kw in TITLE_KEYWORDS):
            trimmed.append(p)
        else:
            # 不含职务关键词，停止追加（去掉后续所有非职务部分）
            break
    title = ','.join(trimmed) if trimmed else title

    # 8. 去每个分段前的行政机构名前缀（保留最终职务）
    #    例："省自然资源厅党组书记" → "党组书记"
    #    例："省局党组书记" → "党组书记"（"局"是机构残留，"党组书记"是完整职务）
    #    例："党组书记" → 保留完整（"党组"是职务名的一部分）
    #    例："省委常委" → 保留完整（无关键词命中，不截断）
    #    规则：
    #      - 找职务关键词位置，若其前有「党组/党委/党工委/纪委」紧接，则从「党组/党委」处截断（保留党组+职务）
    #      - 否则从职务关键词处截断（去掉行政机构前缀）
    #      - 若未找到任何关键词（first_kw_pos == len(p)），保留原文
    PARTY_ORG_WORDS = ('党组', '党委', '党工委', '纪委')
    new_parts = []
    for part in title.split(','):
        p = part.strip()
        if not p:
            continue
        # 找第一个职务关键词的位置
        first_kw_pos = len(p)
        for kw in TITLE_KEYWORDS:
            pos = p.find(kw)
            if pos >= 0:
                first_kw_pos = min(first_kw_pos, pos)
        # 没找到任何关键词 → 保留原文，不截断
        if first_kw_pos < len(p) and first_kw_pos > 0:
            prefix = p[:first_kw_pos]
            # 检查前缀是否以"党组/党委"等结尾（如"省局党组"→从"党组"处截断）
            cut_pos = first_kw_pos
            for party_word in PARTY_ORG_WORDS:
                if prefix.endswith(party_word):
                    # 从"党组/党委"的起始位置截断，保留"党组书记"形式
                    cut_pos = first_kw_pos - len(party_word)
                    break
            p = p[cut_pos:]
        new_parts.append(p)
    title = ','.join(new_parts) if new_parts else title

    return title


def normalize_name(name: str) -> str:
    """规范化姓名：去除空格"""
    if not name:
        return ""
    return name.replace(' ', '').replace('\u3000', '').strip()


def compare_titles(title_excel: str, title_web: str, name: str = "") -> str:
    """
    比对职务是否一致。
    规则：忽略「广东」「广东省」前缀后，职务条目数量相同且一一对应才算一致。
    - 规范化后拆分各职务条目，数量必须相等且内容完全匹配。
    - 不允许网页多出/少于表格中的职务条目（多出兼职也算不一致）。
    """
    t1 = normalize_title(title_excel, name)
    t2 = normalize_title(title_web, name)
    if not t1 or not t2:
        return "不一致"
    if t1 == t2:
        return "一致"
    # 拆分为各职务条目（逗号分隔）
    parts1 = [p.strip() for p in t1.split(',') if p.strip()]
    parts2 = [p.strip() for p in t2.split(',') if p.strip()]
    # 数量不同 → 不一致
    if len(parts1) != len(parts2):
        return "不一致"
    # 数量相同，逐条比对（已经过 normalize_title 去除广东前缀和机构名）
    if parts1 == parts2:
        return "一致"
    return "不一致"


def get_domain(url: str) -> str:
    """提取域名"""
    try:
        parsed = urlparse(url)
        domain = parsed.netloc.lower()
        domain = re.sub(r'^www\.', '', domain)
        return domain
    except Exception:
        return ""


# ─────────────────────────────────────────────
# HTTP 请求
# ─────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def fetch_page(url: str, timeout: int = 15) -> Optional[BeautifulSoup]:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout, verify=False)
        resp.encoding = resp.apparent_encoding or 'utf-8'
        return BeautifulSoup(resp.text, 'html.parser')
    except Exception as e:
        print(f"  [错误] 请求失败 {url}: {e}", file=sys.stderr)
        return None


# ─────────────────────────────────────────────
# 从文本中提取职务
# ─────────────────────────────────────────────

def extract_title_from_text(text: str, norm_name: str) -> str:
    """
    从页面文本中提取现任职务，综合多种策略。
    优先在姓名所在的行/段落内提取，避免列表页抓到其他人的职务。
    """
    if not text:
        return ""

    text_norm = text.replace('\u3000', ' ')
    text_flat = text_norm.replace(' ', '')

    # ── 前置：列表页保护 ──
    # 若姓名在文本中出现，先收集"姓名附近的段落"作为候选文本
    # 这样可以避免列表页里抓到其他人的职务
    name_scoped_text = _scope_text_by_name(text, norm_name)
    scoped_flat = name_scoped_text.replace(' ', '').replace('\u3000', '') if name_scoped_text else ""

    # ── 策略0.5：列表行格式 "{姓名} + 多空格 + {职务}" 或 "{职务} {姓名}" ──
    # 专门处理如"王   曦     省委常委、统战部部长，省政协党组副书记"这类列表页行
    # 在全文各行中找姓名所在行，提取后面的职务部分
    lines_all = [l for l in text.split('\n') if l.strip()]

    def _clean(s: str) -> str:
        """去除各类空白字符（含\xa0不间断空格）"""
        return re.sub(r'[\s\u3000\xa0]+', '', s)

    for line in lines_all:
        line_stripped = line.strip()
        # 检查该行是否包含姓名（含多空格/\xa0变体）
        line_nospace = _clean(line_stripped)
        if norm_name not in line_nospace:
            continue
        # 允许姓名中任意字符之间插入多个空格/全角空格/\xa0（如"王   曦"、"吕 业 升"）
        if len(norm_name) >= 2:
            name_pattern = r'[\s\u3000\xa0]*'.join(re.escape(c) for c in norm_name)
        else:
            name_pattern = re.escape(norm_name)
        m = re.search(name_pattern, line_stripped)
        if not m:
            continue
        # 姓名之后的内容（职务在后）
        after_name = _clean(line_stripped[m.end():])
        # 过滤：after开头是标点/破折号（如"-广东人大网"、"，男，汉族..."）
        if after_name and re.match(r'^[-－，,。；;：:]', after_name):
            after_name = ""
        # 过滤：含"-+网站/门户/官网"的页面标题格式（如"副主任-广东人大网"）
        if after_name and re.search(r'[-－]', after_name) and re.search(r'网|门户|官网', after_name):
            after_name = ""
        # 过滤：裸职务词（≤4字且不含机构名前缀），如"副主任"、"书记"
        # 这类往往来自"{姓名}{职务}"格式的标题行，后文正文会有更完整描述
        if after_name and len(after_name) <= 4 and any(after_name == kw or after_name.endswith(kw) for kw in ['主任','副主任','书记','副书记','省长','副省长','市长','副市长','厅长','副厅长','局长','副局长']):
            after_name = ""
        if (after_name and any(kw in after_name for kw in TITLE_KEYWORDS)
                and 3 <= len(after_name) <= 40):
            return after_name
        # 姓名之前的内容（职务在前，如"党组书记郭跃文"）
        # 要求长度 >= 6 才够完整，避免返回"市长"这类过短导航标签
        before_name = _clean(line_stripped[:m.start()])
        if (before_name and any(kw in before_name for kw in TITLE_KEYWORDS)
                and 6 <= len(before_name) <= 40):
            return before_name

    # ── 策略0.8："{职务}\n{姓名}" 跨行格式（如"党组书记\n郭跃文"）──
    # 要求上一行长度 >= 4，避免误命中导航标签（如"市长"仅作为栏目名）
    # 要求职务行长度 >= 6 才够完整（避免返回"党组书记"这种过短结果）
    for i, line in enumerate(lines_all):
        line_f = _clean(line.strip())
        if line_f == norm_name and i > 0:
            prev_line_f = _clean(lines_all[i-1].strip())
            if any(kw in prev_line_f for kw in TITLE_KEYWORDS) and 6 <= len(prev_line_f) <= 40:
                return prev_line_f

    # ── 策略0.9："{姓名}\n{职务}" 跨行格式（如"张科\n省民族宗教委党组书记、主任"）──
    for i, line in enumerate(lines_all):
        line_f = _clean(line.strip())
        if line_f == norm_name and i < len(lines_all) - 1:
            next_line_f = _clean(lines_all[i+1].strip())
            # 过滤：下一行以姓名开头（如"黄宁生副主任"是标题行，不是独立职务行）
            if next_line_f.startswith(norm_name):
                continue
            if any(kw in next_line_f for kw in TITLE_KEYWORDS) and 4 <= len(next_line_f) <= 50:
                return next_line_f

    # ── 策略1："现任XXX" 关键词（优先，可获取完整多职务字符串）──
    # 先在 scoped 文本找，再回退到全文
    # 注意：过滤"现任领导"这类导航词（要求提取内容长度>4字）
    for search_flat in ([scoped_flat, text_flat] if scoped_flat else [text_flat]):
        pattern_current = re.compile(r'现任([^。\n\r]{4,120})')
        for m in pattern_current.finditer(search_flat):
            candidate = m.group(1).strip()
            if any(kw in candidate for kw in TITLE_KEYWORDS) and len(candidate) > 4:
                candidate = re.split(r'[。]', candidate)[0].strip()
                return candidate
        if search_flat == scoped_flat and not scoped_flat:
            break

    # ── 策略2："担任XXX" 关键词 ──
    for search_flat in ([scoped_flat, text_flat] if scoped_flat else [text_flat]):
        pattern_serve = re.compile(r'担任([^。\n\r]{2,120})')
        for m in pattern_serve.finditer(search_flat):
            candidate = m.group(1).strip()
            if any(kw in candidate for kw in TITLE_KEYWORDS):
                candidate = re.split(r'[。]', candidate)[0].strip()
                return candidate
        if search_flat == scoped_flat and not scoped_flat:
            break

    # ── 策略3："{职务} {姓名}" 格式（如"省长　孟凡利"）──
    name_variants = [norm_name]
    if len(norm_name) == 2:
        name_variants.append(norm_name[0] + ' ' + norm_name[1])
        name_variants.append(norm_name[0] + '\u3000' + norm_name[1])

    search_text = name_scoped_text if name_scoped_text else text_norm
    for nv in name_variants:
        pattern = r'([^\n\r，,。；;]{2,30}?)\s*' + re.escape(nv) + r'(?:\s|$|，|。)'
        for m in re.finditer(pattern, search_text):
            candidate = m.group(1).strip()
            candidate = re.sub(r'\s+', '', candidate)
            if any(kw in candidate for kw in TITLE_KEYWORDS) and 2 <= len(candidate) <= 40:
                return candidate

    # ── 策略4：姓名前后窗口内找职务关键词短语 ──
    name_pos = text_flat.find(norm_name)
    if name_pos >= 0:
        window_start = max(0, name_pos - 150)
        window_end = min(len(text_flat), name_pos + len(norm_name) + 150)
        window = text_flat[window_start:window_end]

        candidates = []
        for kw in TITLE_KEYWORDS:
            kw_pos = window.find(kw)
            if kw_pos < 0:
                continue
            phrase_start = 0
            for sep in ['，', ',', '\n', '\r', '。', '：', ':', '；', ';']:
                idx = window.rfind(sep, 0, kw_pos)
                if idx >= 0:
                    phrase_start = max(phrase_start, idx + 1)
            phrase_end = len(window)
            for sep in ['，', ',', '\n', '\r', '。', '；', ';']:
                idx = window.find(sep, kw_pos + len(kw))
                if idx >= 0:
                    phrase_end = min(phrase_end, idx)
            phrase = window[phrase_start:phrase_end].strip()
            if 2 <= len(phrase) <= 50:
                dist = abs(kw_pos - (name_pos - window_start))
                candidates.append((dist, phrase))

        if candidates:
            candidates.sort(key=lambda x: x[0])
            return candidates[0][1]

    # ── 策略5：全文中出现的职务关键词行（过滤页面标题行）──
    # 优先找正文中的职务行（从后往前，但跳过无关行）
    _FOOTER_PATTERNS = re.compile(r'主办单位|版权所有|网站标识码|ICP备|公网安备|联系我们|网站声明|承办单位')
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    for line in reversed(lines):
        line_flat = _clean(line)
        if not (any(kw in line_flat for kw in TITLE_KEYWORDS) and 2 <= len(line_flat) <= 60):
            continue
        # 过滤：含"姓名-单位"格式的页面标题行（如"张科-广东省民族宗教事务委员会网站"）
        if re.search(r'[\-－]', line_flat) and ('网站' in line_flat or '门户' in line_flat or '官网' in line_flat):
            continue
        # 过滤：页脚/版权行（含"主办单位"、"版权所有"等）
        if _FOOTER_PATTERNS.search(line_flat):
            continue
        # 过滤：纯导航词（"现任领导"、"主任会议"等，长度<=4且不含人名）
        if len(line_flat) <= 4 and norm_name not in line_flat:
            continue
        return line_flat

    return ""


def _scope_text_by_name(text: str, norm_name: str) -> str:
    """
    列表页保护：在文本中找到姓名所在的行或段落，
    返回以该位置为中心的500字符窗口文本。
    若文本较短（<800字，通常是个人详情页）则直接返回空串（用全文）。
    """
    # 短文本（个人详情页）不做收窄
    clean = text.replace(' ', '').replace('\u3000', '')
    if len(clean) < 1000:
        return ""

    pos = clean.find(norm_name)
    if pos < 0:
        # 尝试两字姓名中间加空格的变体
        if len(norm_name) == 2:
            for sep in [' ', '\u3000']:
                variant = norm_name[0] + sep + norm_name[1]
                pos2 = text.find(variant)
                if pos2 >= 0:
                    start = max(0, pos2 - 200)
                    end = min(len(text), pos2 + 300)
                    return text[start:end]
        return ""

    # 返回姓名前200字符+后300字符的窗口（覆盖列表行格式）
    start = max(0, pos - 200)
    end = min(len(clean), pos + 300)
    return clean[start:end]


# ─────────────────────────────────────────────
# 各网站专项策略
# ─────────────────────────────────────────────

def extract_from_gd_gov(soup: BeautifulSoup, name: str) -> str:
    """
    省政府 gd.gov.cn：
    - 有"省长　孟凡利"这种职务+姓名格式
    - 简历中含"现任"关键词
    """
    norm_name = normalize_name(name)
    full_text = soup.get_text(separator='\n', strip=True)
    return extract_title_from_text(full_text, norm_name)


def extract_from_gdszx(soup: BeautifulSoup, name: str) -> str:
    """
    省政协 gdszx.gov.cn：
    职务信息在"个人简历"板块中，通常最后一条是现任职务。
    """
    norm_name = normalize_name(name)

    # 定位"个人简历"标题
    resume_tag = None
    for tag in soup.find_all(True):
        if '个人简历' in (tag.get_text() or ''):
            # 找最小的包含该文本的节点
            if len(tag.get_text()) < 20:
                resume_tag = tag
                break

    if resume_tag:
        # 取后续兄弟节点
        siblings_text = []
        for sib in resume_tag.find_next_siblings():
            t = sib.get_text(separator='\n', strip=True)
            if t:
                siblings_text.append(t)
        if siblings_text:
            bio_text = '\n'.join(siblings_text)
            result = extract_title_from_text(bio_text, norm_name)
            if result:
                return result

        # 取父节点全文
        parent = resume_tag.find_parent()
        if parent:
            result = extract_title_from_text(parent.get_text(separator='\n', strip=True), norm_name)
            if result:
                return result

    full_text = soup.get_text(separator='\n', strip=True)
    return extract_title_from_text(full_text, norm_name)


def extract_from_gdpc(soup: BeautifulSoup, name: str) -> str:
    """
    广东人大 gdpc.gov.cn：
    职务信息在头像下方描述或正文简介中。
    页面格式通常：{职务}副主任\n时间\n来源\n{姓名}，男，...，现任省人大常委会XXX。
    """
    norm_name = normalize_name(name)
    full_text = soup.get_text(separator='\n', strip=True)
    return extract_title_from_text(full_text, norm_name)


def extract_generic(soup: BeautifulSoup, name: str) -> str:
    """通用策略"""
    norm_name = normalize_name(name)
    full_text = soup.get_text(separator='\n', strip=True)
    return extract_title_from_text(full_text, norm_name)


def extract_title_from_url(url: str, soup: BeautifulSoup, name: str) -> str:
    """根据域名选择策略"""
    domain = get_domain(url)
    if 'gd.gov.cn' in domain:
        return extract_from_gd_gov(soup, name)
    elif 'gdszx.gov.cn' in domain:
        return extract_from_gdszx(soup, name)
    elif 'gdpc.gov.cn' in domain:
        return extract_from_gdpc(soup, name)
    else:
        return extract_generic(soup, name)


# ─────────────────────────────────────────────
# 主逻辑
# ─────────────────────────────────────────────

def process_excel(filepath: str, delay: float = 1.5, start_row: int = 2, limit: int = 0):
    """
    处理Excel文件。
    列映射：
      A(0)=姓名, B(1)=机构路径, C(2)=粤政易职务, D(3)=网上公开职务,
      E(4)=网址, F(5)=巡检结果(写入), G(6)=网页职务(写入)
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 确保 G 列有表头
    if ws.cell(row=1, column=7).value is None:
        ws.cell(row=1, column=7).value = '网页爬取职务'
        wb.save(filepath)

    total = ws.max_row - 1
    print(f"[INFO] 共 {total} 条数据（排除首行表头），从第{start_row}行开始处理")
    if limit > 0:
        print(f"[INFO] 限制处理 {limit} 条")
    print("-" * 70)

    processed = 0
    for row_idx in range(start_row, ws.max_row + 1):
        name = str(ws.cell(row=row_idx, column=1).value or '').strip()       # A列
        org_path = str(ws.cell(row=row_idx, column=2).value or '').strip()   # B列
        title_system = str(ws.cell(row=row_idx, column=3).value or '').strip() # C列
        title_excel = str(ws.cell(row=row_idx, column=4).value or '').strip() # D列
        url = str(ws.cell(row=row_idx, column=5).value or '').strip()         # E列

        f_cell = ws.cell(row=row_idx, column=6)   # F列：巡检结果
        g_cell = ws.cell(row=row_idx, column=7)   # G列：网页职务

        if not name:
            continue

        if limit > 0 and processed >= limit:
            break

        print(f"[行{row_idx}] {name} | D列: {title_excel}")

        # 条件1：无账号（B列含"暂无粤政易账号"）
        if '暂无粤政易账号' in org_path:
            f_cell.value = '无帐号'
            g_cell.value = ''
            print(f"  → 无帐号，跳过")
            wb.save(filepath)
            processed += 1
            continue

        # 条件2：无网址
        _NO_URL_PATTERNS = re.compile(r'^(无|None|nan|无官网|管理员确认|待管理员|暂无|无官方|无参考)', re.IGNORECASE)
        if not url or url in ('None', 'nan', '') or _NO_URL_PATTERNS.match(url) or not url.startswith('http'):
            f_cell.value = '无网址'
            g_cell.value = ''
            print(f"  → 无网址，跳过")
            wb.save(filepath)
            processed += 1
            continue

        # 爬取网页
        print(f"  → 爬取: {url}")
        soup = fetch_page(url)

        if soup is None:
            f_cell.value = '不一致'
            g_cell.value = '[页面加载失败]'
            print(f"  → 页面加载失败")
            wb.save(filepath)
            processed += 1
            time.sleep(delay)
            continue

        title_web = extract_title_from_url(url, soup, name)

        if title_web:
            title_web = title_web.strip()
            result = compare_titles(title_excel, title_web, name)
        else:
            title_web = '[未找到职务信息]'
            result = '不一致'

        f_cell.value = result
        g_cell.value = title_web

        print(f"  → 网页职务: {title_web}")
        print(f"  → D列职务:  {title_excel}")
        print(f"  → 比对结果: {result}")

        wb.save(filepath)
        processed += 1
        time.sleep(delay + random.uniform(0, 0.5))

    wb.save(filepath)
    print("-" * 70)
    print(f"[完成] 共处理 {processed} 条，结果已写入: {filepath}")


# ─────────────────────────────────────────────
# 命令行入口
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='重点用户职务信息巡检工具',
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument('file', help='Excel文件路径（.xlsx）')
    parser.add_argument('--delay', type=float, default=1.5,
                        help='每次请求间隔秒数（默认1.5秒）')
    parser.add_argument('--start', type=int, default=2,
                        help='从第几行开始处理（默认2，跳过表头）')
    parser.add_argument('--limit', type=int, default=0,
                        help='最多处理多少条（默认0=全部）')
    args = parser.parse_args()

    process_excel(args.file, delay=args.delay, start_row=args.start, limit=args.limit)


if __name__ == '__main__':
    main()
